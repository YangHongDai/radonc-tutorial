"""Extract Estes 2025 trials Excel into structured JSON."""
import openpyxl, json, os, re

SRC = r"/Users/arthurdai/Desktop/radonc-tutorial/Rad Onc Tables - Key Studies Estes 2025.xlsx"
OUT = r"/Users/arthurdai/Desktop/radonc-tutorial/radonc-tutorial-main/data/trials.json"

FIELDS = {
    8: "topic", 9: "subtopic", 10: "trial", 11: "citation",
    12: "n", 13: "patients", 14: "methods", 15: "results",
    16: "interpretation", 17: "commentary", 18: "notes",
    19: "references", 20: "commentary_refs", 25: "flag",
    26: "question", 27: "answer",
}

SHEETS = ["Breast", "CNS", "Cutaneous", "GI", "GU", "Gyn", "H&N",
          "Lung", "Lymphoma", "Palliation", "Peds", "Prostate", "Sarcoma", "Benign"]

def clean(v):
    if v is None: return ""
    s = str(v).strip()
    s = re.sub(r"\s+", " ", s)
    return s

def compact(v, maxlen=800):
    s = clean(v)
    if len(s) > maxlen:
        s = s[:maxlen] + "…"
    return s

wb = openpyxl.load_workbook(SRC, data_only=True)
out = {}
total = 0
for sheet in SHEETS:
    if sheet not in wb.sheetnames: continue
    ws = wb[sheet]
    trials = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or all(c is None or str(c).strip() == "" for c in row[:20]):
            continue
        trial = clean(row[10]) if len(row) > 10 else ""
        topic = clean(row[8]) if len(row) > 8 else ""
        if not trial and not topic: continue
        rec = {}
        for idx, key in FIELDS.items():
            if idx < len(row):
                v = row[idx]
                if key in ("methods", "results", "interpretation", "commentary", "notes", "patients"):
                    rec[key] = compact(v, 600)
                else:
                    rec[key] = clean(v)
        # skip section header rows (no citation, no trial with content)
        if not rec.get("trial") and not rec.get("citation"):
            continue
        trials.append(rec)
    out[sheet] = trials
    total += len(trials)
    print(f"{sheet}: {len(trials)} trials")

os.makedirs(os.path.dirname(OUT), exist_ok=True)
with open(OUT, "w", encoding="utf-8") as f:
    json.dump(out, f, ensure_ascii=False, indent=1)
print(f"Total: {total} trials -> {OUT}")
